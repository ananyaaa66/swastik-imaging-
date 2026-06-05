"""
Contact management routes.

GET    /api/contacts         – list / search contacts (admin)
POST   /api/contacts         – create single contact (admin)
POST   /api/contacts/upload  – CSV/Excel bulk import (admin)
GET    /api/contacts/count   – total contact count (admin)
DELETE /api/contacts/{id}    – delete contact (admin)
"""

from __future__ import annotations

import csv
import io
import re
from datetime import datetime, timezone
from typing import Optional

from bson import ObjectId
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from pydantic import BaseModel, Field

from app.auth import get_current_admin
from app.database import get_db

router = APIRouter(prefix="/api/contacts", tags=["contacts"])

_PHONE_RE = re.compile(r"^\d{10}$")


# ── Pydantic models ──────────────────────────────────────────────────────


class ContactCreateRequest(BaseModel):
    patientName: str = Field(..., min_length=1)
    phone: str = Field(..., pattern=r"^\d{10}$")
    email: Optional[str] = ""
    source: str = "manual-import"


# ── Helper ────────────────────────────────────────────────────────────────


def _serialize(doc: dict) -> dict:
    doc["id"] = str(doc.pop("_id"))
    return doc


def _clean_phone(raw: str) -> str:
    """Extract last 10 digits from a phone string."""
    val = str(raw).strip()
    if val.endswith(".0"):
        val = val[:-2]
    digits = re.sub(r"\D", "", val)
    if len(digits) >= 10:
        return digits[-10:]
    return digits



# ── Routes ────────────────────────────────────────────────────────────────


@router.get("/count")
async def contact_count(_admin: dict = Depends(get_current_admin)):
    db = get_db()
    total = await db.contacts.count_documents({})
    return {"total": total}


@router.get("")
async def list_contacts(
    search: Optional[str] = Query(None),
    limit: int = Query(200, ge=1, le=1000),
    _admin: dict = Depends(get_current_admin),
):
    db = get_db()
    query: dict = {}
    if search:
        search_regex = {"$regex": search, "$options": "i"}
        query["$or"] = [
            {"patientName": search_regex},
            {"phone": search_regex},
            {"email": search_regex},
        ]

    cursor = db.contacts.find(query).sort("createdAt", -1).limit(limit)
    results = []
    async for doc in cursor:
        results.append(_serialize(doc))
    return results


@router.post("", status_code=201)
async def create_contact(
    body: ContactCreateRequest,
    _admin: dict = Depends(get_current_admin),
):
    db = get_db()
    existing = await db.contacts.find_one({"phone": body.phone})
    if existing:
        raise HTTPException(status_code=409, detail="Contact with this phone already exists")

    doc = {
        "patientName": body.patientName,
        "phone": body.phone,
        "email": body.email or "",
        "source": body.source,
        "createdAt": datetime.now(timezone.utc).isoformat(),
    }
    result = await db.contacts.insert_one(doc)
    doc["id"] = str(result.inserted_id)
    doc.pop("_id", None)
    return doc


@router.post("/upload")
async def upload_contacts(
    file: UploadFile = File(...),
    _admin: dict = Depends(get_current_admin),
):
    """Import contacts from CSV or Excel file.

    Expected columns (case-insensitive, flexible naming):
      - Name / Patient Name / Full Name / patientName
      - Phone / Phone Number / Mobile / phoneNumber / Contact
      - Email (optional)

    Returns: { total, imported, skipped, failed, errors[] }
    """
    if not file.filename:
        raise HTTPException(status_code=400, detail="No file provided")

    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    if ext not in ("csv", "xlsx", "xls"):
        raise HTTPException(
            status_code=400,
            detail="Unsupported file type. Please upload .csv or .xlsx",
        )

    raw = await file.read()
    rows: list[dict] = []

    if ext == "csv":
        text = raw.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        for row in reader:
            rows.append({k.strip().lower(): (v or "").strip() for k, v in row.items() if k})
    else:
        try:
            import openpyxl
        except ImportError:
            raise HTTPException(status_code=500, detail="openpyxl not installed on server")

        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            raise HTTPException(status_code=400, detail="Excel file has no sheets")

        header_row = None
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            header_row = [str(c or "").strip().lower() for c in row]
            break

        if not header_row:
            raise HTTPException(status_code=400, detail="Excel file has no header row")

        for row in ws.iter_rows(min_row=2, values_only=True):
            record = {}
            for i, val in enumerate(row):
                if i < len(header_row):
                    record[header_row[i]] = str(val or "").strip()
            if any(record.values()):
                rows.append(record)
        wb.close()

    # Column name mapping
    name_keys = ["name", "patient name", "full name", "patientname", "patient_name"]
    phone_keys = ["phone", "phone number", "mobile", "phonenumber", "phone_number", "contact", "mobile number"]
    email_keys = ["email", "email address", "e-mail", "emailaddress"]

    def _find(row: dict, keys: list[str]) -> str:
        for k in keys:
            if k in row and row[k]:
                return row[k]
        return ""

    db = get_db()
    now = datetime.now(timezone.utc).isoformat()
    total = len(rows)
    imported = 0
    skipped = 0
    errors: list[dict] = []

    for idx, row in enumerate(rows, start=2):  # start=2 because row 1 is header
        name = _find(row, name_keys)
        phone_raw = _find(row, phone_keys)
        email = _find(row, email_keys)

        phone = _clean_phone(phone_raw)

        if not phone or not _PHONE_RE.match(phone):
            errors.append({"row": idx, "error": f"Invalid phone number: '{phone_raw}'"})
            continue

        if not name:
            name = "Unknown"

        try:
            result = await db.contacts.update_one(
                {"phone": phone},
                {
                    "$setOnInsert": {
                        "patientName": name,
                        "phone": phone,
                        "email": email,
                        "source": "manual-import",
                        "createdAt": now,
                    }
                },
                upsert=True,
            )
            if result.upserted_id:
                imported += 1
            else:
                skipped += 1
        except Exception as exc:
            errors.append({"row": idx, "error": str(exc)})

    return {
        "total": total,
        "imported": imported,
        "skipped": skipped,
        "failed": errors[:50],  # Return array of errors to match frontend UploadResult interface
    }



@router.delete("/{contact_id}", status_code=204)
async def delete_contact(
    contact_id: str,
    _admin: dict = Depends(get_current_admin),
):
    db = get_db()
    try:
        result = await db.contacts.delete_one({"_id": ObjectId(contact_id)})
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid contact ID")
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Contact not found")
